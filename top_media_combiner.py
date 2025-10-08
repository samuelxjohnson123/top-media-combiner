# top_media_combiner.py  — Streamlit app (single file)
# Features:
# - Combine Sprinklr + Cision
# - Resolve URLs (keeps full MSN permalinks)
# - Fill missing Journalist for Forbes / TechRadar / Tom’s Guide (CSS + JSON-LD + AMP fallback)
# - Map Group/Outlet via explicit rules + master list
# - Flag duplicates, ExUS authors, REMOVE, Non-US URL, MSN Non-US → ? = 'R'
# - NEW: "R Reason" (final column) explains why 'R' was applied
# - Journalist column appears immediately before Sentiment
# - Journalist names auto-filled by the app are colored BLUE in the Excel export
# - Styled Excel output with clickable full URLs

import streamlit as st
import pandas as pd
from io import BytesIO
import os
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import re
from typing import Optional
from urllib.parse import urlparse, urlunparse
import tldextract
from bs4 import BeautifulSoup

st.title("📊 Top Media Combiner")

st.markdown("""
Upload your daily **Sprinklr** and **Cision** files. This app will:

✅ Combine & align columns  
✅ Resolve programmatic URLs to final destination for deduplication  
✅ Show resolved URLs in the output for manual checking  
✅ Map `Group` & `Outlet` from URL patterns or master list (including domains)  
✅ Retain proper casing from master list  
✅ Mark duplicates (only later ones) with `R` in `?`  
✅ Fill missing **Journalist** for Forbes/TechRadar/Tom’s Guide, then apply **ExUS** mapping  
✅ Flag non-US URL editions (subdomain and first-path-segment on known outlets) with `R`  
✅ Make URLs clickable, showing full resolved URL text  
✅ Format dates as `m/d/yyyy`  
✅ Add blank columns for manual entry  
✅ Show live progress bar during URL resolution  
✅ Styled header + column widths  
✅ `Source Platform` column (`Sprinklr` or `Cision`) at the end  
✅ **NEW**: `R Reason` (last column) explains why a row was flagged `R`
""")

# ---------------------------
# Config & Inputs
# ---------------------------
sprinklr_file = st.file_uploader("Upload Sprinklr file (.xlsx or .csv)", type=["xlsx", "csv"])
cision_file   = st.file_uploader("Upload Cision file (.xlsx or .csv)",   type=["xlsx", "csv"])

MASTER_FILE = "2025_Master Outlet List.xlsx"
if not os.path.exists(MASTER_FILE):
    st.error(f"Master outlet list file `{MASTER_FILE}` not found in app folder.")
    st.stop()

master_xl = pd.ExcelFile(MASTER_FILE)

# ---------------------------
# MSN handling helpers
# ---------------------------
_US_MSN_LOCALES = {"en-us", "es-us"}  # locales we KEEP; everything else is REMOVE

def _msn_locale(path: str) -> Optional[str]:
    """Extracts first path segment that looks like a locale (e.g., 'en-us', 'en-gb')."""
    m = re.match(r"/([a-z]{2}-[a-z]{2})(?:/|$)", path.lower())
    return m.group(1) if m else None

def resolve_url(url):
    """
    MSN links: keep the original permalink to avoid collapsing to generic msn.com.
    Others: follow redirects via GET (stream=True to avoid full body download).
    """
    if pd.isna(url) or url == "":
        return ""
    try:
        if "msn.com" in str(url).lower():
            return url  # preserve full MSN permalink
        r = requests.get(url, allow_redirects=True, timeout=6, stream=True, headers={"User-Agent": "Mozilla/5.0"})
        return r.url
    except:
        return url

# ---------------------------
# Non-US locale detection
# ---------------------------
NON_US_SUBDOMAIN_TOKENS = {
    "uk","gb","au","me","sea","ca","sg","hk","nz","ie","eu","mena","latam","sa","za",
    "de","fr","es","it","nl","se","no","dk","fi","pl","tr",
    "jp","kr","cn","tw","vn","th","ph","my","id",
    "ae","qa","kw","bh",
    "br","mx","ar","cl","co","pe",
    "apac","emea"
}

# Only apply *path* locale detection to these to avoid false positives like linkedin.com/in/...
PATH_LOCALE_ALLOWED_DOMAINS = {
    "techradar.com","www.techradar.com",
    "tomsguide.com","www.tomsguide.com",
    "gamesradar.com","www.gamesradar.com",
    "androidcentral.com","www.androidcentral.com",
    "windowscentral.com","www.windowscentral.com",
    "laptopmag.com","www.laptopmag.com",
    "imore.com","www.imore.com",
    "ign.com","www.ign.com",
    "cnet.com","www.cnet.com",
    # requested additions:
    "forbes.com","www.forbes.com",
    "yahoo.com","www.yahoo.com",
    "bbc.com","www.bbc.com",
    "engadget.com","www.engadget.com",
    "digitaltrends.com","www.digitaltrends.com",
}

PATH_LOCALE_TOKENS = set(NON_US_SUBDOMAIN_TOKENS)  # reuse the same tokens for first path segment
_lang_region_re = re.compile(r"^[a-z]{2,3}[-_](?:[a-z]{2}|[a-z]{3})$", re.I)  # en-gb, pt-br, zh-hk, en_US

def _base_domain_and_subtokens(host: str):
    host = (host or "").lower()
    ext = tldextract.extract(host)
    base = ".".join([p for p in [ext.domain, ext.suffix] if p])
    subs = [t for t in ext.subdomain.split(".") if t]
    return base, subs

def mark_non_us_url(u: str) -> bool:
    """Return True if URL is clearly non-US (subdomain token or allow-listed first path segment)."""
    if not u:
        return False
    try:
        p = urlparse(str(u))
        base, subs = _base_domain_and_subtokens(p.netloc)

        # Rule 1: subdomain tokens like me.mashable.com, uk.yahoo.com
        if any(tok in NON_US_SUBDOMAIN_TOKENS for tok in subs):
            return True

        # Rule 2: known outlets using first path segment locale (/uk/, /pt-br/, /zh-hk/)
        if base in PATH_LOCALE_ALLOWED_DOMAINS:
            first_seg = next((seg for seg in p.path.split("/") if seg), "").lower()
            if first_seg in PATH_LOCALE_TOKENS or _lang_region_re.match(first_seg):
                # Do not flag explicit US variants
                if first_seg.endswith("us") or first_seg in {"us","en-us","en_us"}:
                    return False
                return True
    except Exception:
        return False
    return False

# ---------------------------
# Journalist scraping (Forbes/TechRadar/Tom’s Guide)
# ---------------------------
TARGET_SCRAPE_DOMAINS = {
    "forbes.com","www.forbes.com",
    "techradar.com","www.techradar.com",
    "tomsguide.com","www.tomsguide.com",
}

UA = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                   "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"),
    "Accept-Language": "en-US,en;q=0.8",
}

def _fetch_html(url: str) -> str:
    try:
        r = requests.get(url, timeout=10, headers=UA, allow_redirects=True)
        if r.status_code < 400:
            return r.text or ""
    except Exception:
        pass
    return ""

def _extract_author_from_html(domain: str, html: str) -> Optional[str]:
    soup = BeautifulSoup(html, "html.parser")
    d = (domain or "").lower()

    # Forbes
    if "forbes" in d:
        meta = soup.find("meta", attrs={"name": "author"})
        if meta and meta.get("content"):
            return re.sub(r"^by\s+", "", meta["content"].strip(), flags=re.I)
        tag = soup.select_one("a[rel='author'], span.contributor__name")
        if tag:
            return re.sub(r"^by\s+", "", tag.get_text(strip=True), flags=re.I)

    # TechRadar (Future plc sites) — selectors + meta + JSON-LD
    if "techradar" in d:
        for sel in [
            "span.by-author",
            "span.byline__name",
            ".article-byline__author",
            "a.author",
            "a[rel='author']",
            "a[href*='/author/']",
            "a[itemprop='url'] span[itemprop='name']",
            "[itemprop='author'] [itemprop='name']",
        ]:
            tag = soup.select_one(sel)
            if tag:
                txt = tag.get_text(strip=True)
                if txt:
                    return re.sub(r"^by\s+", "", txt, flags=re.I)

        meta = soup.find("meta", attrs={"name": "author"})
        if meta and meta.get("content"):
            return re.sub(r"^by\s+", "", meta["content"].strip(), flags=re.I)

        for s in soup.find_all("script", attrs={"type": "application/ld+json"}):
            try:
                import json
                data = s.string
                if not data:
                    continue
                obj = json.loads(data)
                items = obj if isinstance(obj, list) else [obj]
                names = []
                for it in items:
                    t = str(it.get("@type", "")).lower()
                    if t in ("article", "newsarticle", "blogposting"):
                        auth = it.get("author")
                        if isinstance(auth, dict) and auth.get("name"):
                            names.append(str(auth["name"]).strip())
                        elif isinstance(auth, list):
                            for a in auth:
                                if isinstance(a, dict) and a.get("name"):
                                    names.append(str(a["name"]).strip())
                if names:
                    return "; ".join([re.sub(r"^by\s+", "", n, flags=re.I) for n in names if n])
            except Exception:
                continue

    # Tom's Guide
    if "tomsguide" in d:
        tag = soup.select_one("span.author-name, a[rel='author'], a[href*='/author/']")
        if tag:
            return re.sub(r"^by\s+", "", tag.get_text(strip=True), flags=re.I)
        meta = soup.find("meta", attrs={"name": "author"})
        if meta and meta.get("content"):
            return re.sub(r"^by\s+", "", meta["content"].strip(), flags=re.I)

    return None

def fill_missing_journalist(permalink: str, journalist: str) -> str:
    """If Journalist blank and domain is target, fetch & parse byline (with AMP fallback)."""
    if pd.notna(journalist) and str(journalist).strip():
        return journalist
    try:
        host = urlparse(permalink).netloc.lower()
        if host not in TARGET_SCRAPE_DOMAINS:
            return journalist

        # Try canonical page
        html = _fetch_html(permalink)
        name = _extract_author_from_html(host, html) if html else None
        if name:
            return name

        # AMP fallback (very consistent on TR)
        amp_url = permalink.rstrip("/") + "/amp"
        html_amp = _fetch_html(amp_url)
        name = _extract_author_from_html(host, html_amp) if html_amp else None
        if name:
            return name
    except Exception:
        pass
    return journalist

# ---------------------------
# Group/Outlet mapping
# ---------------------------
def map_group_outlet(url, pub_name, master_map):
    """
    Maps Group/Outlet via explicit rules first, then falls back to master_map.
    - MSN: keep only /en-us and /es-us; all other locales or no-locale => REMOVE
    - Yahoo: map to vertical-specific brands
    """
    url_lc = str(url).strip().lower()
    pub_name_lc = str(pub_name).strip().lower()

    # --- MSN handling ---
    if "msn.com" in url_lc:
        # Extract path portion safely
        try:
            path = re.sub(r"^[a-z]+://[^/]+", "", url_lc)  # strip scheme+host
        except:
            path = url_lc
        loc = _msn_locale(path)
        if loc in _US_MSN_LOCALES:
            return "Tech", "MSN"
        else:
            return "REMOVE", "REMOVE"

    # --- Yahoo handling ---
    if "sports.yahoo" in url_lc:
        return "Lifestyle", "Yahoo! Sports"
    if "yahoo.com/entertainment" in url_lc:
        return "Lifestyle", "Yahoo! Entertainment"
    if "yahoo.com/lifestyle" in url_lc:
        return "Lifestyle", "Yahoo! Lifestyle"
    if "finance.yahoo.com" in url_lc:
        return "Tech", "Yahoo! Finance"
    if "yahoo.com/news" in url_lc:
        return "Tech", "Yahoo! News"
    if "yahoo.com/tech" in url_lc:
        return "Tech", "Yahoo! Tech"

    # --- Master list fallback ---
    for key in (pub_name_lc, url_lc):
        if key in master_map:
            return master_map[key]['Group'], master_map[key]['Outlet']
    return "", ""

def extract_cision_url(cell):
    if pd.isna(cell):
        return ""
    match = re.search(r'HYPERLINK\("([^"]+)"', str(cell))
    return match.group(1) if match else str(cell)

# ---------------------------
# Main App
# ---------------------------
if sprinklr_file and cision_file:
    # Load inputs
    if sprinklr_file.name.endswith(".csv"):
        sprinklr = pd.read_csv(sprinklr_file, encoding='utf-8', errors='ignore')
    else:
        sprinklr = pd.read_excel(sprinklr_file)

    if cision_file.name.endswith(".csv"):
        cision = pd.read_csv(cision_file, skiprows=3, encoding='utf-8')
    else:
        cision = pd.read_excel(cision_file, skiprows=3)

    cision.dropna(how='all', inplace=True)

    if sprinklr.empty or cision.empty:
        st.error("One of the uploaded files appears to be empty or malformed.")
        st.stop()

    sprinklr.columns = sprinklr.columns.str.strip()
    cision.columns   = cision.columns.str.strip()

    sprinklr['Source Platform'] = 'Sprinklr'
    cision['Source Platform']   = 'Cision'

    detailed_list    = master_xl.parse("Detailed List for Msmt")
    journalist_check = master_xl.parse("Journalist Check")

    # Keep Sprinklr's own Media Title, only rename Resolved_URL
    rename_map = {"Resolved_URL": "Permalink"}
    sprinklr = sprinklr.rename(columns=rename_map)

    # Drop Conversation Stream if it exists
    if "Conversation Stream" in sprinklr.columns:
        sprinklr = sprinklr.drop(columns=["Conversation Stream"])

    cision = cision.rename(columns={
        "Date": "CreatedTime",
        "Media Type": "Source",
        "Media Outlet": "Publication Name",
        "Title": "Media Title",
        "Link": "Permalink",
        "Author": "Journalist",
        "Sentiment": "Sentiment"
    })
    cision['Permalink'] = cision['Permalink'].apply(extract_cision_url)

    sprinklr = sprinklr.loc[:, ~sprinklr.columns.duplicated()]
    cision   = cision.loc[:, ~cision.columns.duplicated()]

    common_cols = [
        "CreatedTime", "Source", "Publication Name", "Media Title",
        "Permalink", "Journalist", "Sentiment", "Source Platform"
    ]
    sprinklr = sprinklr.reindex(columns=common_cols, fill_value="")
    cision   = cision.reindex(columns=common_cols,   fill_value="")

    sprinklr['Publication Name'] = sprinklr['Publication Name'].str.strip()
    cision['Publication Name']   = cision['Publication Name'].str.strip()
    detailed_list['Outlet Name'] = detailed_list['Outlet Name'].str.strip()

    combined = pd.concat([sprinklr, cision], ignore_index=True)

    # Date formatting
    combined['CreatedTime'] = pd.to_datetime(combined['CreatedTime'], errors='coerce').dt.strftime('%-m/%-d/%Y')

    # Add blank columns for downstream manual entry
    for col in ['?', 'Campaign', 'Phase', 'Products', 'PreOrder', 'Group', 'Outlet', 'ExUS Author']:
        if col not in combined.columns:
            combined[col] = ""

    # Build master outlet map
    master_map = {}
    for _, row in detailed_list.iterrows():
        group = row['Vertical (FOR VLOOKUP)']
        outlet = row['Outlet Name']
        for key in [
            str(row['Outlet Name']).strip().lower(),
            str(row['Outlet Name From Searches']).strip().lower(),
            str(row['URL']).strip().lower()
        ]:
            if key and key != 'nan':
                master_map[key] = {'Group': group, 'Outlet': outlet}

    # ExUS author marking (from "Journalist Check" sheet)
    journalist_check['key'] = journalist_check['Publication'].str.lower().str.strip() + "|" + journalist_check['Name'].str.lower().str.strip()
    combined['key'] = combined['Publication Name'].str.lower().str.strip() + "|" + combined['Journalist'].str.lower().str.strip()
    exus_keys = set(journalist_check[journalist_check['Geo'].str.upper() == 'EXUS']['key'])
    combined['ExUS Author'] = combined['key'].apply(lambda x: 'Yes' if x in exus_keys else '')
    combined.drop(columns=['key'], inplace=True)

    # Resolve URLs + mapping + byline fill + non-US flag
    st.info("🔄 Resolving final URLs, filling journalists, and mapping…")
    progress_bar = st.progress(0, text="Resolving URLs… 0%")
    status_text = st.empty()

    resolved_urls   = []
    groups          = []
    outlets         = []
    nonus_flags     = []
    filled_journos  = []
    journalist_filled_flags = []
    n_rows = len(combined)

    for idx, row in combined.iterrows():
        resolved = resolve_url(row['Permalink'])
        resolved_urls.append(resolved)

        # Fill missing journalist (Forbes/TechRadar/Tom's Guide) with AMP fallback
        orig_j = row.get('Journalist', '')
        j_filled = fill_missing_journalist(resolved, orig_j)
        filled_journos.append(j_filled)
        journalist_filled_flags.append(
            (not str(orig_j).strip()) and bool(str(j_filled).strip())
        )

        # Group/Outlet mapping
        group, outlet = map_group_outlet(resolved, row['Publication Name'], master_map)
        groups.append(group)
        outlets.append(outlet)

        # Non-US URL flag
        nonus_flags.append('R' if mark_non_us_url(resolved) else '')

        pct_complete = (idx + 1) / n_rows
        progress_bar.progress(pct_complete, text=f"Processing… {idx+1}/{n_rows}")
        status_text.text(f"Processed {idx+1} of {n_rows}")

    status_text.text("✅ URL resolution, journalist fill, and mapping complete.")
    combined['Resolved_Permalink'] = resolved_urls
    combined['Group']     = groups
    combined['Outlet']    = outlets
    combined['Journalist']= filled_journos  # (now updated)

    # Re-run ExUS author mapping now that some bylines were filled
    combined['key'] = combined['Publication Name'].str.lower().str.strip() + "|" + combined['Journalist'].str.lower().str.strip()
    combined['ExUS Author'] = combined['key'].apply(lambda x: 'Yes' if x in exus_keys else '')
    combined.drop(columns=['key'], inplace=True)

    # ---------------------------
    # Centralized ? = 'R' logic + R Reason (final column)
    # ---------------------------
    combined['?'] = ''

    # a) duplicates (by resolved URL)
    combined['Resolved_Permalink_lower'] = combined['Resolved_Permalink'].str.lower()
    mask_with_url = combined['Resolved_Permalink_lower'].notna() & (combined['Resolved_Permalink_lower'] != "")
    dup_mask = mask_with_url & combined.duplicated(subset=['Resolved_Permalink_lower'], keep='first')

    # b) ExUS Author
    exus_mask = combined['ExUS Author'].eq('Yes')

    # c) REMOVE group/outlet
    remove_group_mask  = combined['Group'].str.upper().eq('REMOVE')
    remove_outlet_mask = combined['Outlet'].str.upper().eq('REMOVE')

    # d) Non-US URL patterns (from loop)
    nonus_mask = pd.Series(nonus_flags, index=combined.index).eq('R')

    # e) MSN non-US (explicit)
    path_series = combined['Resolved_Permalink'].fillna('').str.lower().str.replace(r'^[a-z]+://[^/]+', '', regex=True)
    msn_mask = combined['Resolved_Permalink'].str.contains('msn.com', case=False, na=False)
    msn_loc  = path_series.apply(_msn_locale)
    msn_non_us_mask = msn_mask & ~msn_loc.isin(_US_MSN_LOCALES)

    # Final '?' flag
    combined.loc[
        dup_mask | exus_mask | remove_group_mask | remove_outlet_mask | nonus_mask | msn_non_us_mask,
        '?'
    ] = 'R'

    # Build R Reason strings (ordered, multiple allowed)
    reasons = [[] for _ in range(len(combined))]
    def _add_reason(mask, label):
        for i in combined.index[mask]:
            reasons[i].append(label)

    _add_reason(dup_mask,             'Duplicate URL')
    _add_reason(exus_mask,            'ExUS Author')
    _add_reason(remove_group_mask,    'REMOVE Group')
    _add_reason(remove_outlet_mask,   'REMOVE Outlet')
    _add_reason(msn_non_us_mask,      'MSN Non-US')
    _add_reason(nonus_mask & ~msn_non_us_mask, 'Non-US URL')

    combined['R Reason'] = ['; '.join(r) for r in reasons]

    # Clean up helper
    combined.drop(columns=['Resolved_Permalink_lower'], inplace=True)

    # Show final URL as clickable hyperlink (display the full URL)
    combined['Permalink'] = combined['Resolved_Permalink']
    combined.drop(columns=['Resolved_Permalink'], inplace=True)
    combined['Permalink'] = combined['Permalink'].apply(
        lambda x: f'=HYPERLINK("{x}", "{x}")' if pd.notna(x) and x != "" else ""
    )

    # Final column ordering (Journalist immediately before Sentiment)
    final_cols = [
        'CreatedTime', 'Source', 'Publication Name', 'Group', 'Outlet',
        'Media Title', 'Permalink', '?', 'Campaign', 'Phase', 'Products', 'PreOrder',
        'Journalist', 'Sentiment',
        'Country', 'Total News Media Potential Reach', 'Web shares overall', 'EMV',
        'ExUS Author', 'Source Platform',
        'R Reason',  # <-- last column
    ]
    for col in final_cols:
        if col not in combined.columns:
            combined[col] = ""
    combined = combined[final_cols]

    # ---------------------------
    # Write Excel with styles
    # ---------------------------
    out = BytesIO()
    combined.to_excel(out, index=False, engine='openpyxl')
    out.seek(0)

    wb = load_workbook(out)
    ws = wb.active
    ws.auto_filter.ref = ws.dimensions

    header_font = Font(bold=True, color="0000F5")
    header_alignment = Alignment(horizontal='center')
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment

    col_widths = {
        "Media Title": 40.5,
        "Permalink": 40.5,
        "Outlet": 18
    }
    header_map = {cell.value: cell.column for cell in ws[1]}
    for col_name, width in col_widths.items():
        if col_name in header_map:
            col_letter = get_column_letter(header_map[col_name])
            ws.column_dimensions[col_letter].width = width

    if "Permalink" in header_map:
        permalink_col_letter = get_column_letter(header_map["Permalink"])
        for row in ws.iter_rows(
            min_row=2,
            min_col=ws[permalink_col_letter+'1'].column,
            max_col=ws[permalink_col_letter+'1'].column
        ):
            for cell in row:
                cell.alignment = Alignment(
                    wrap_text=False,
                    horizontal='left',
                    shrink_to_fit=False
                )

    # Color Journalist cells blue if the app populated them
    if "Journalist" in header_map:
        j_col_idx = header_map["Journalist"]  # 1-based
        j_col_letter = get_column_letter(j_col_idx)
        for i, filled in enumerate(journalist_filled_flags, start=2):  # data starts at row 2
            if filled:
                ws[f"{j_col_letter}{i}"].font = Font(color="0000F5")

    styled_out = BytesIO()
    wb.save(styled_out)
    styled_out.seek(0)

    st.success("✅ Processing complete! Download your combined file below:")
    st.download_button(
        label="📥 Download Combined File",
        data=styled_out.getvalue(),
        file_name="Top_Media_Combined.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

else:
    st.info("⬆️ Please upload both files to begin.")
